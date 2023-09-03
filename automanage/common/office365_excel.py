#!/usr/bin/env python3
# -*- coding-utf-8 -*-

import os
import time
import glob
import fire
from loguru import logger

import csv
import openpyxl
import pandas as pd

class ExcelUtils(object):
    """
    处理 Excel 文件。

    Example:
        python3 office365.py main
    """
    def __init__(self):
        pass

    def all_excel(self, excel_path, excel_type = 'xlsx'):
        all_excel = excel_path + '*.' + excel_type
        excel_list = glob.glob(all_excel)
        if excel_list:
            logger.debug(f"Excel list: {excel_list}")
        else:
            excel_list = []
            logger.error(f"Error: No type '{excel_type}' file found in '{excel_path}'.")
        return excel_list

    def create_xlsx_by_openpyxl(self, xlsx_path, sheet_name = 'sheet1'):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = sheet_name
        # wb.remove(sheet)
        wb.save(xlsx_path)
    
    def create_sheet_by_openpyxl(self, xlsx_path, sheets_name = ['sheet1']):
        self.excel_exists(self.create_xlsx_by_openpyxl, xlsx_path)
        wb = openpyxl.load_workbook(xlsx_path)
        for sheet_name in sheets_name:
            if sheet_name in wb.sheetnames:
                logger.debug(f"Sheet '{sheet_name}' already exists in '{xlsx_path}'.")
                continue
            wb.create_sheet(sheet_name)
            logger.debug(f"Create sheet '{sheet_name}' in '{xlsx_path}'.")
        wb.save(xlsx_path)

    def remove_sheet_by_openpyxl(self, xlsx_path, sheets_name = ['sheet1']):
        self.excel_exists(self.create_xlsx_by_openpyxl, xlsx_path)
        wb = openpyxl.load_workbook(xlsx_path)
        for sheet_name in sheets_name:
            if sheet_name not in wb.sheetnames:
                logger.debug(f"Sheet '{sheet_name}' not found in '{xlsx_path}'.")
                continue
            ws = wb[sheet_name]
            wb.remove(ws)
            logger.debug(f"Remove sheet '{sheet_name}' in '{xlsx_path}'.")
        wb.save(xlsx_path)

    def create_csv_by_csv(self, csv_path):
        # csv 库无多个sheet 概念
        with open(csv_path, 'w', newline = '', encoding = 'utf-8') as csvfile:
            pass
        return
    
    def excel_exists(self, func, excel_path):
        if not os.path.isfile(excel_path):
            logger.error(f"File '{excel_path}' not found.")
            logger.info(f"Start to create '{excel_path}'.")
            func(excel_path)
            logger.info(f"Finished to create '{excel_path}'.")

    def excel_exception(self, excel_path):
        try:
            df = pd.read_csv(excel_path, encoding = 'gbk', thousands = ',', index_col = False)
        except UnicodeDecodeError:
            df = pd.read_csv(excel_path, encoding = 'utf-8', thousands = ',', index_col = False)
        except UnicodeDecodeError:
            data = pd.read_csv(excel_path, encoding = 'ANSI', thousands = ',', index_col = False)
        except UnicodeDecodeError:
            data = pd.read_csv(excel_path, encoding = 'gb2312', thousands = ',', index_col = False)
        except Exception as e:
            logger.error(f"Failed to read CSV file '{excel_path}': {e}")
            return None, False
        return df, True

    def convert_csv_to_xlsx(self, csv_path, xlsx_path, xlsx_sheet_name = 'sheet1'):
        """
        将 CSV 文件转换为 XLSX 文件。
        Args:
            csv_path (str): CSV 文件的路径。
            xlsx_path (str): XLSX 文件的路径。
            xlsx_sheet_name (list, optional): 需要转化的 sheet 名称。缺省值为 sheet1。若存在则覆盖，不存在则创建。
        Returns:
            None
        """
        if not os.path.isfile(csv_path):
            logger.error(f"File '{csv_path}' not found.")
            return
        self.excel_exists(self.create_xlsx_by_openpyxl, xlsx_path)
        self.create_sheet_by_openpyxl(xlsx_path)
        logger.info(f"Start to convert '{csv_path}' to '{xlsx_path}'.")
        df, sign = self.excel_exception(csv_path)
        if not sign:
            return
        with pd.ExcelWriter(xlsx_path, mode='a', engine='openpyxl', if_sheet_exists = 'replace') as writer:
            df.to_excel(writer, sheet_name = xlsx_sheet_name, index = False)
        logger.info(f"Finished to convert '{csv_path}' to '{xlsx_path}'.")
        self.remove_sheet_by_openpyxl(xlsx_path)

    def merge_xlsx_to_xlsx(self, multi_xlsx_path, xlsx_path, xlsx_sheet_name = 'sheet1'): # 单 sheet，默认 merge 到 sheet1
        for xlsx_file in multi_xlsx_path:
            if not os.path.isfile(xlsx_file):
                logger.error(f"File '{xlsx_file}' not found.")
                continue
            logger.info(f"Start to merge '{xlsx_file}' to '{xlsx_path}'.")
            # wb = openpyxl.load_workbook(xlsx_file)
            # sheet_names = wb.sheetnames
            # for sheet_name in sheet_names:
            #     logger.debug(f"Start to merge sheet '{sheet_name}' to '{xlsx_path}', '{xlsx_sheet_name}'.")
            #     df = pd.read_excel(xlsx_file, sheet_name = sheet_name)
            #     with pd.ExcelWriter(xlsx_path, mode='a', engine='openpyxl') as writer:
            #         df.to_excel(writer, sheet_name = xlsx_sheet_name, index = False)
            # logger.info(f"Finished to merge '{xlsx_file}' to '{xlsx_path}'.")
            # wb.save(xlsx_file)

    def merge_xlsx_to_xlsx_sheets(self, multi_xlsx_path, xlsx_path, xlsx_sheets_name = ['sheet1'], sign = True): # 多sheet
        self.excel_exists(self.create_xlsx_by_openpyxl, xlsx_path)
     
    def merge_csv_multi_to_single(self, multi_csv, single_csv):
        self.excel_exists(self.create_csv_by_csv, single_csv)
        df_list = []
        for csv_file in multi_csv:
            if not os.path.isfile(csv_file):
                logger.error(f"File '{csv_file}' not found.")
                continue
            logger.info(f"Start to merge '{csv_file}' to '{single_csv}'.")
            df, sign = self.excel_exception(csv_file)
            if not sign:
                continue
            df_list.append(df)
        if df_list:
            df_merged = pd.concat(df_list, ignore_index=True)
            with pd.ExcelWriter(single_csv, mode='a', engine='openpyxl') as writer:
                df_merged.to_excel(writer, index=False, encoding='utf-8')
            logger.info(f"Finished to merge '{csv_file}' to '{single_csv}'.")

    def merge_csv_to_xlsx(self, multi_csv, xlsx_path, xlsx_sheet_name = 'sheet1'): # 单 sheet
        temp_csv = 'temp.csv'
        logger.info(f"Start to merge '{multi_csv}' to '{xlsx_path}'.")
        self.merge_csv_multi_to_single(multi_csv, temp_csv)
        self.convert_csv_to_xlsx(temp_csv, xlsx_path, xlsx_sheet_name)
        logger.info(f'Delete temp file: {temp_csv}')
        os.remove(temp_csv)
        logger.info(f"Finished to merge '{multi_csv}' to '{xlsx_path}'.")
    
    def merge_csv_to_xlsx_sheets(self, multi_csv, xlsx_path, xlsx_sheets_name = ['sheet1'], sign = True): # sign 为 True 多 sheet，False 单 sheet
        if sign:
            if len(multi_csv) < len(xlsx_sheets_name):
                xlsx_sheets_name = xlsx_sheets_name[0:len(multi_csv)]
            else:
                extra_sheets_name = [f'sheet{i}' for i in range(len(xlsx_sheets_name), len(multi_csv))]
                xlsx_sheets_name = xlsx_sheets_name + extra_sheets_name
            self.excel_exists(self.create_xlsx_by_openpyxl, xlsx_path)
            self.create_sheet_by_openpyxl(xlsx_path, xlsx_sheets_name)
        else:
            sheet_name = xlsx_sheets_name
        for csv_file in multi_csv:
            if not os.path.isfile(csv_file):
                logger.error(f"File '{csv_file}' not found.")
                continue
            if sign:
                sheet_name = xlsx_sheets_name[multi_csv.index(csv_file)] # 通过索引获取 sheet name
            logger.info(f"Start to merge '{csv_file}' to '{xlsx_path}' with sheet name '{sheet_name}'.")
            df, sign = self.excel_exception(csv_file)
            if not sign:
                continue
            with pd.ExcelWriter(xlsx_path, mode='a', engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name = sheet_name, index = False)
            logger.info(f"Finished to merge '{csv_file}' to '{xlsx_path}' with sheet name '{sheet_name}'.")
        logger.info(f"Finished to merge '{multi_csv}' to '{xlsx_path}'.")


    def read_csv(self, csv_path):
        """
        读取 CSV 文件并返回一个包含所有行的列表。
        Args:
            csv_path (str): CSV 文件的路径。
        Returns:
            list: 包含所有行的列表。
        """
        sheet = []
        try:
            with open(csv_path, newline = '') as csvfile:
                reader = csv.reader(csvfile, delimiter = ',', quotechar = '"')
                for row in reader:
                    sheet.append(row)
        except csv.Error as e:
            logger.error(f"Error: Failed to read CSV file '{csv_path}': {e}")
        return sheet



    def show_csv_rows(self, sheet, start_row = 0, end_row = None):
        if not sheet:
            logger.warning("Warning: Empty CSV file.")
            return
        if start_row < 0 or start_row >= len(sheet):
            logger.warning(f"Warning: Invalid start row index '{start_row}'.")
            return
        sub_sheet = sheet[start_row:end_row] if end_row is not None else sheet[start_row]
        if end_row is not None and (end_row < 0 or end_row >= len(sheet)):
            logger.warning(f"Warning: Invalid end row index '{end_row}'.")
            return
        for row in sub_sheet:
            logger.debug(row)
        return sub_sheet

    def show_csv_columns(self, sheet, start_col = 0, end_col = None):
        """
        显示 CSV 文件 start_col 到 end_col 列内容。
        Args:
            sheet (list): 包含所有行的列表。
            start_col (int): 起始列的索引。
            end_col (int): 结束列的索引。
        Returns:
            sub_sheet: 返回一个包含 start_col 到 end_col 列的列表。
        """
        if not sheet:
            logger.warning("Warning: Empty CSV file.")
            return
        if start_col < 0 or start_col >= len(sheet[0]):
            logger.warning(f"Warning: Invalid start column index '{start_col}'.")
            return
        sub_sheet = [row[start_col:end_col] if end_col is not None else row[start_col] for row in sheet]
        if end_col is not None and (end_col < 0 or end_col >= len(sheet[0])):
            logger.warning(f"Warning: Invalid end column index '{end_col}'.")
            return
        for row in sub_sheet:
            logger.debug(row)
        return sub_sheet



    def main(self):
        multi_xlsx = ['1.xlsx', '2.xlsx', '3.xlsx']
        self.merge_xlsx_to_xlsx(multi_xlsx, 'a.xlsx')

        excel_name = 'f5_big_ip_assets_data'
        # self.convert_csv_to_xlsx(excel_name + '.csv', excel_name + '.xlsx', ['sh1'])

        multi_csv = ['1.csv', '2.csv', '3.csv']
        single_csv = 'a.csv'
        # self.merge_csv_multi_to_single(multi_csv, single_csv)

        sheets_name = ['sh1', 'sh2', 'sh3']
        # for multi in multi_csv:
        #     sheet = multi.replace('.csv', '.xlsx')
        #     self.convert_csv_to_xlsx(multi, sheet, ['sh1'])

        # for multi in multi_csv:
        #     sheet = multi.replace('.csv', '.xlsx')
        #     self.create_sheet_by_openpyxl(sheet, sheets_name)




if __name__ == '__main__':
    fire.Fire(ExcelUtils)
